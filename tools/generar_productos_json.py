#!/usr/bin/env python3
"""
Genera data/productos.json desde el Excel de ubicación de farmacia.

Uso:
  python tools/generar_productos_json.py
  python tools/generar_productos_json.py "UBICACION MEDICAMENTOS FARMACIA 2026 (1).xlsx"

Requiere:
  pip install openpyxl
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = ["Producto", "MUEBLE", "ID_MUEBLE", "UBICACION"]


def norm_space(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).strip())


def slug_id(value: str) -> str:
    value = norm_space(value).upper()
    value = re.sub(r"[^A-Z0-9]+", "_", value).strip("_")
    return value or "SIN_ID"


def find_default_excel(repo_root: Path) -> Path:
    candidates = [
        p for p in repo_root.glob("*.xlsx")
        if not p.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError("No se encontró ningún .xlsx en la raíz del repositorio.")
    if len(candidates) > 1:
        named = [p for p in candidates if "UBICACION" in p.name.upper()]
        if named:
            return sorted(named)[0]
        raise RuntimeError(
            "Hay más de un .xlsx en la raíz. Indique el archivo explícitamente: "
            "python tools/generar_productos_json.py archivo.xlsx"
        )
    return candidates[0]


def read_planograma_map(wb) -> dict[str, str]:
    if "planograma" not in wb.sheetnames:
        return {}
    ws = wb["planograma"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [norm_space(v).upper() for v in rows[0]]
    try:
        name_idx = header.index("NOMBRE MUEBLE")
        id_idx = header.index("ID")
    except ValueError:
        return {}

    mapping: dict[str, str] = {}
    for row in rows[1:]:
        name = norm_space(row[name_idx] if name_idx < len(row) else "")
        code = norm_space(row[id_idx] if id_idx < len(row) else "")
        if name and code:
            mapping[name.upper()] = code
    return mapping


def locate_header(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        row = [norm_space(c.value) for c in ws[row_idx]]
        header_positions: dict[str, int] = {}
        for required in REQUIRED_COLUMNS:
            if required in row:
                header_positions[required] = row.index(required)
        if len(header_positions) == len(REQUIRED_COLUMNS):
            return row_idx, header_positions

    raise RuntimeError(
        "No se encontraron las columnas requeridas: "
        + ", ".join(REQUIRED_COLUMNS)
    )


def build_productos(excel_path: Path, sheet_name: str = "ubicacion") -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"No existe la hoja '{sheet_name}'. Hojas disponibles: {', '.join(wb.sheetnames)}")

    planograma_map = read_planograma_map(wb)
    ws = wb[sheet_name]
    header_row, pos = locate_header(ws)

    productos: list[dict[str, str]] = []
    warnings: list[dict[str, Any]] = []

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        producto = norm_space(row[pos["Producto"]] if pos["Producto"] < len(row) else "")
        if not producto:
            continue

        mueble = norm_space(row[pos["MUEBLE"]] if pos["MUEBLE"] < len(row) else "")
        id_mueble = norm_space(row[pos["ID_MUEBLE"]] if pos["ID_MUEBLE"] < len(row) else "")
        ubicacion = norm_space(row[pos["UBICACION"]] if pos["UBICACION"] < len(row) else "")

        if not id_mueble:
            id_mueble = planograma_map.get(mueble.upper(), "") or slug_id(mueble)
            warnings.append({
                "fila": excel_row_number,
                "tipo": "ID_MUEBLE vacío",
                "mueble": mueble,
                "id_mueble_generado": id_mueble,
            })

        productos.append({
            "producto": producto,
            "mueble": mueble,
            "id_mueble": id_mueble,
            "ubicacion": ubicacion,
        })

    return productos, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description="Genera data/productos.json desde Excel.")
    parser.add_argument("excel", nargs="?", help="Ruta al archivo Excel. Si se omite, busca un .xlsx en la raíz.")
    parser.add_argument("--sheet", default="ubicacion", help="Nombre de hoja con la tabla principal.")
    parser.add_argument("--output", default="data/productos.json", help="Salida JSON.")
    parser.add_argument("--report", default="data/reporte_excel.json", help="Reporte de generación.")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    excel_path = Path(args.excel) if args.excel else find_default_excel(repo_root)
    if not excel_path.is_absolute():
        excel_path = repo_root / excel_path
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    productos, warnings = build_productos(excel_path, args.sheet)

    output_path = repo_root / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(productos, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    report = {
        "excel": excel_path.name,
        "total_productos": len(productos),
        "warnings": warnings,
        "campos": ["producto", "mueble", "id_mueble", "ubicacion"],
    }
    report_path = repo_root / args.report
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"OK: {len(productos)} productos exportados a {output_path.relative_to(repo_root)}")
    if warnings:
        print(f"Advertencias: {len(warnings)} filas con ID_MUEBLE vacío fueron normalizadas.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
